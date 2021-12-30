import http.client
import ssl
import base64
import string
import json
import copy

from phpipam_client import PhpIpamClient, GET, PATCH, POST

from ipaddress import IPv4Interface
from ipaddress import IPv6Interface

from ipaddress import ip_network
from ipaddress import IPv6Network

from ipaddress import ip_address
from ipaddress import IPv6Address

import ipaddress

import requests
import logging

from configparser import ConfigParser

# Disable requests' warnings for insecure connections
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


def GetFortiApiToken(ipaddr, username, password, port):
  timeout=10
  urlbase = "https://{ipaddr}:{port}/".format(ipaddr=ipaddr,port=port)
  
  session = requests.session()
  url = urlbase + 'logincheck'
  session.post(url,data='username={username}&secretkey={password}'.format(username=username,password=password),verify=False)
  for cookie in session.cookies:
    if cookie.name == 'ccsrftoken':
      csrftoken = cookie.value[1:-1]  # strip quotes
      session.headers.update({'X-CSRFTOKEN': csrftoken})
  
  #login_check = session.get(urlbase + "api/v2/monitor/router/ipv4/?vdom=DOP_DGWAN")
  #login_check.raise_for_status()
  return session

def GetFortiApi(token, uri, vdom):
  request = token.get(uri, verify=False, timeout=10, params='vdom='+vdom)
  if request.status_code == 200:
    return request.json()['results']
  else:
    return request.status_code

def FortiDefaultv6(bluevlanID, cust):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'AnycastGateway':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip']
  cfgblue = """
      edit 0
        set gateway %s
        set device "cust%s_red1"
        set comment "Extern"
    next
  """ % (blueipv6, nummer)
  return cfgblue

def FortiDefaultv4(bluevlanID, cust):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'AnycastGateway':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip']
  cfgblue = """
      edit 0
        set gateway %s
        set device "cust%s_red1"
        set comment "Extern"
    next
  """ % (blueipv4, nummer)
  return cfgblue

def FortiRoutev6(bluevlanID, cust):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'AnycastGateway':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip']
  cfgblue = """
      edit 0
        set dst 2a0c:ed80::/29
        set gateway %s
        set device "cust%s_blue1"
    next
  """ % (blueipv6, nummer)
  return cfgblue 

def FortiRoutev4(bluevlanID, cust):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'AnycastGateway':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip']
  cfgblue = """
      edit 0
        set dst 185.213.35.0 255.255.255.0
        set gateway %s
        set device "cust%s_blue1"
    next
  """ % (blueipv4, nummer)
  cfgblue2 = """
      edit 0
        set dst 100.64.247.0 255.255.255.0
        set gateway %s
        set device "cust%s_blue1"
    next
  """ % (blueipv4, nummer)
  cfgblue  = cfgblue + cfgblue2
  return cfgblue

def CLIGreen(bluevlanID, cust):
  """Erstellen CLI fuer Fortimanager

  Args:
      bluevlanID (int): IPAM Vlan ID
      cust (str): customer123 Bezeichnung

  Returns:
      str: Fortigate CLI fuer Blue
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip'] + "/" + b['mask']

  cfgblue = """
      edit "cust%s_green1"
      set alias "cust%s_green-Server"
          set vdom "%s"
          set status down
        set ip %s
        set allowaccess ping
          config ipv6
              set ip6-address %s
              set ip6-allowaccess ping
              end
          set interface "Port-Channel11"
          set vlanid %s
      next
  """ % (nummer, nummer, cust, blueipv4, blueipv6, intblue)
  return cfgblue

def CLIBlue(bluevlanID, cust):
  """Erstellen CLI fuer Fortimanager

  Args:
      bluevlanID (int): IPAM Vlan ID
      cust (str): customer123 Bezeichnung

  Returns:
      str: Fortigate CLI fuer Blue
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip'] + "/" + b['mask']

  cfgblue = """
      edit "cust%s_blue1"
        set alias "cust%s_blue-Service"
          set vdom "%s"
          set status down
        set ip %s
        set allowaccess ping
          config ipv6
              set ip6-address %s
              set ip6-allowaccess ping
              end
          set interface "Port-Channel12"
          set vlanid %s
      next
  """ % (nummer, nummer, cust, blueipv4, blueipv6, intblue)
  return cfgblue

def CLIOrange(bluevlanID, cust):
  """Erstellen CLI fuer Fortimanager

  Args:
      bluevlanID (int): IPAM Vlan ID
      cust (str): customer123 Bezeichnung

  Returns:
      str: Fortigate CLI fuer Blue
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip'] + "/" + b['mask']

  cfgorange = """
    edit "cust%s_orange1"
      set alias "cust%s_orange-Public"
        set vdom "%s"
        set status down
      set ip %s
      set allowaccess ping
        config ipv6
            set ip6-address %s
            set ip6-allowaccess ping
            end
        set interface "Port-Channel14"
        set vlanid %s
    next
    """ % (nummer, nummer, cust, blueipv4, blueipv6, intblue)
  return cfgorange

def CLIRed(bluevlanID, cust):
  """Erstellen CLI fuer Fortimanager

  Args:
      bluevlanID (int): IPAM Vlan ID
      cust (str): customer123 Bezeichnung

  Returns:
      str: Fortigate CLI fuer Blue
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']
  nummer = cust.replace('customer','')

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  blueipv4 = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  blueipv6 = add['ip'] + "/" + b['mask']

  cfgred = """
  edit "cust%s_red1"
      set alias "cust%s_red-Extern"
      set vdom "%s"
      set status down
    set ip %s
    set allowaccess ping
      config ipv6
          set ip6-address %s
          set ip6-allowaccess ping
          end
      set interface "Port-Channel13"
      set vlanid %s
  next
  """ % (nummer, nummer, cust, blueipv4, blueipv6, intblue)
  return cfgred

def CreateNewSection(name, cust):
  """Erstellen neuer Section in Ipam

  Args:
      name (str): Kunden Name (z. B. Dokumental)
      cust (str): Firewall vdom Name (z. B. customer023)

  Returns:
      str: Returnt neue Section id
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  secname = name + " IPv4"
  secdesc = cust + '_' + name
  IPAMvlans = ipam.post('/sections/', {
      'name': secname, 
      'description': secdesc, 
      'masterSection': '4', 
      'permissions': '{"2":"2","3":"1","4":"3"}', 
      'strictMode': '1', 
      'subnetOrdering': 'default', 
      'order': None, 
      'showVLAN': '1', 
      'showVRF': '1', 
      'showSupernetOnly': '1', 
      'DNS': None
  })

  searchstring = "/sections/"
  sectionid = ipam.get(searchstring, {
      'filter_by': 'name',
      'filter_value': secname,
  })

  return sectionid[0]['id']

def NewGreenSubnet():
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  for start in range(75, 100):
      searchstring = "/subnets/search/10.113." + str(start) + ".0/24/"
      bluev6Info = ipam.get(searchstring)
      if bluev6Info == 0:
          return start

def DCNMv6Route(vrf, vlanidred, subnetid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  bluevlanID = vlanidred
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  v4ipred = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  v6ipred = add['ip']

  querystring = "/subnets/" + subnetid + "/"
  blue = ipam.get(querystring)
  v6iporange = blue['subnet'] + "/" + blue['mask']

  scriptstring = "& python c:/Temp/Git/Cisco/DCNM/new-route6.py --v %s --p %s --n %s --r %s --t %s" % (vrf, v6iporange, v6ipred, description, "FIXME")
  return scriptstring

def DCNMv4Route(vrf, vlanidred, vlanidorange, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  bluevlanID = vlanidred
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  v4ipred = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  v6ipred = add['ip']
  
  bluevlanID = vlanidorange
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  v4iporange = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  v6iporange = add['ip'] + "/" + b['mask']


  scriptstring = "& python c:/Temp/Git/Cisco/DCNM/new-route.py --v %s --p %s --n %s --r %s --t %s" % (vrf, v4iporange, v4ipred, description, "FIXME")
  return scriptstring

def ForcepointNew(customerID, kuerzel, vlanidorange, customerslash56):
  
  namelang = customerID + "-" + kuerzel
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )

  querystring = "/vlan/" + vlanidorange + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'Fortigate':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  v4iporange = add['ip']
              if isinstance(test, ipaddress.IPv6Address):
                  v6iorange = add['ip']

  querystring = "/subnets/" + customerslash56 + "/"
  blue = ipam.get(querystring)
  v6customer = blue['subnet'] + "/" + blue['mask']
  scriptstring = "& python c:/Temp/Git/Forcepoint/new-customer.py --c %s --n %s --p %s" % (namelang, v4iporange, v6customer)
  return scriptstring

def DCNML3VLAN(vrf, vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
        url='https://ipam.consinto.com',
        app_id='network',
        username=ipamuser,
        ssl_verify=False,
        password=ipampassword,
        user_agent='myapiclient', # custom user-agent header
  )
  bluevlanID = vlanid
  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'vlanId',
      'filter_value': bluevlanID,
  })

  intblue = IPAMvlans[0]['number']

  querystring = "/vlan/" + bluevlanID + "/subnets/"
  blue = ipam.get(querystring)
  for b in blue:
      idb = b['id']
      querystring = "/subnets/" + idb + "/addresses/"
      bb = ipam.get(querystring)
      for add in bb:
          if add['hostname'] == 'AnycastGateway':
              test = ipaddress.ip_address(add['ip'])
              if isinstance(test, ipaddress.IPv4Address):
                  v4ip = add['ip'] + "/" + b['mask']
              if isinstance(test, ipaddress.IPv6Address):
                  v6ip = add['ip'] + "/" + b['mask']


  scriptstring = "& python c:/Temp/Git/Cisco/DCNM/new-l3vlan.py --v %s --i %s --n %s --a %s --b %s" % (vrf, intblue, description, v4ip, v6ip)
  return scriptstring

def CreateGreenSubnetv4(vlanid, network, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )
  subnet = "10.113." + str(network) + ".0"
  IPAMsubnets = "10.113." + str(network) + ".0/24"
  IPAMvlans = ipam.post('/subnets/', {
      'subnet': subnet,
      'mask': 24,
      'sectionId': '1',
      'description': description,
      'vlanId': vlanid
  })

  searchstring = '/subnets/cidr/' + IPAMsubnets + '/'
  redv4Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 1,
  })
  netid = redv4Info[0]['id']
  searchstring = '/addresses/first_free/' + netid + '/'

  IPAMvlans = ipam.post(searchstring, {
      'description': 'Fortigate',
      'hostname': 'Fortigate',
      'deviceId': 84
  })
  return netid

def CreateBlueSubnetv4(vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMsubnets = ipam.get('/subnets/138/first_subnet/31/')

  subnet = IPAMsubnets.split("/")
  IPAMvlans = ipam.post('/subnets/', {
      'subnet': subnet[0],
      'mask': subnet[1],
      'sectionId': '1',
      'description': description,
      'masterSubnetId': 138,
      'vlanId': vlanid
  })

  searchstring = '/subnets/cidr/' + IPAMsubnets + '/'
  redv4Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 1,
  })
  netid = redv4Info[0]['id']
  searchstring = '/addresses/first_free/' + netid + '/'
  IPAMvlans = ipam.post(searchstring, {
      'description': 'AnycastGateway',
      'hostname': 'AnycastGateway',
      'deviceId': 83
  })
  IPAMvlans = ipam.post(searchstring, {
      'description': 'Fortigate',
      'hostname': 'Fortigate',
      'deviceId': 84
  })
  return netid

def CreateRedSubnetv4(vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMsubnets = ipam.get('/subnets/921/first_subnet/31/')

  subnet = IPAMsubnets.split("/")
  IPAMvlans = ipam.post('/subnets/', {
      'subnet': subnet[0],
      'mask': subnet[1],
      'sectionId': '1',
      'description': description,
      'masterSubnetId': 921,
      'vlanId': vlanid
  })

  searchstring = '/subnets/cidr/' + IPAMsubnets + '/'
  redv4Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 1,
  })
  netid = redv4Info[0]['id']
  searchstring = '/addresses/first_free/' + netid + '/'
  IPAMvlans = ipam.post(searchstring, {
      'description': 'AnycastGateway',
      'hostname': 'AnycastGateway',
      'deviceId': 83
  })
  IPAMvlans = ipam.post(searchstring, {
      'description': 'Fortigate',
      'hostname': 'Fortigate',
      'deviceId': 84
  })
  return netid

def CreateRedSubnetv6(vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMv6subnets = ipam.get('/subnets/915/first_subnet/64/')
  v6subnet = IPAMv6subnets.split("/")
  # Das neue Netz wird angelegt
  IPAMvlans = ipam.post('/subnets/915/first_subnet/64/', {
      'description': description,
      'vlanId': vlanid
  })
  # Wir suchen uns die ID des Netzwerks
  searchstring = '/subnets/cidr/' + IPAMv6subnets + '/'
  bluev6Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 2,
  })
  bluev6Info = json.dumps(bluev6Info[0])
  bluev6Info = json.loads(bluev6Info)
  bluev6Id = bluev6Info['id']
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
      'description': 'NA',
      'hostname': 'NA'
  })
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
    'description': 'AnycastGateway',
    'hostname': 'AnycastGateway',
    'deviceId': 83
  })
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
    'description': 'Fortigate',
    'hostname': 'Fortigate',
    'deviceId': 84
  })
  return bluev6Id

def CreateBlueSubnetv6(vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMv6subnets = ipam.get('/subnets/13/first_subnet/127/')
  v6subnet = IPAMv6subnets.split("/")
  # Das neue Netz wird angelegt
  IPAMvlans = ipam.post('/subnets/13/first_subnet/127/', {
      'description': description,
      'vlanId': vlanid
  })
  # Wir suchen uns die ID des Netzwerks
  searchstring = '/subnets/cidr/' + IPAMv6subnets + '/'
  bluev6Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 2,
  })
  bluev6Info = json.dumps(bluev6Info[0])
  bluev6Info = json.loads(bluev6Info)
  bluev6Id = bluev6Info['id']
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
    'description': 'AnycastGateway',
    'hostname': 'AnycastGateway',
    'deviceId': 83
  })
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
    'description': 'Fortigate',
    'hostname': 'Fortigate',
    'deviceId': 84
  })
  return bluev6Id

def CreateOrangeSubnetv4(vlanid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMsubnets = ipam.get('/subnets/920/first_subnet/32/')

  subnet = IPAMsubnets.split("/")

  IPAMvlans = ipam.post('/subnets/', {
      'subnet': subnet[0],
      'mask': subnet[1],
      'sectionId': '1',
      'description': description,
      'masterSubnetId': 920,
      'vlanId': vlanid
  })

  searchstring = '/subnets/cidr/' + IPAMsubnets + '/'
  redv4Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 1,
  })
  netid = redv4Info[0]['id']
  searchstring = '/addresses/first_free/' + netid + '/'
  IPAMvlans = ipam.post(searchstring, {
      'description': 'Fortigate',
      'hostname': 'Fortigate',
      'deviceId': 84
  })
  return netid

def CreateOrangeSubnetv6(vlanid, subnetid, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )
  IPAMv6subnets = ipam.get('/subnets/'+ str(subnetid) + '/first_subnet/64/')
  IPAMvlans = ipam.post('/subnets/' + str(subnetid) + '/first_subnet/64/', {
      'description': description,
      'vlanId': vlanid
  })
  searchstring = '/subnets/cidr/' + IPAMv6subnets + '/'
  bluev6Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 2,
  })
  bluev6Info = json.dumps(bluev6Info[0])
  bluev6Info = json.loads(bluev6Info)
  bluev6Id = bluev6Info['id']
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
      'description': 'NA',
      'hostname': 'NA'
  })
  searchstring = '/addresses/first_free/' + bluev6Id + '/'
  IPAMvlans = ipam.post(searchstring, {
    'description': 'Fortigate',
    'hostname': 'Fortigate',
    'deviceId': 84
  })
  return bluev6Id

def CreateGreenVlan(vlanid, name, description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipamurl = config.get('IPAM', 'url')
  ipampassword = config.get('IPAM', 'ipampassword')


  ipam = PhpIpamClient(
    url=ipamurl,
    app_id='network',
    username=ipamuser,
    ssl_verify=False,
    password=ipampassword,
    user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.post('/vlan/', {
    'domainId': 3,
    'name': name,
    'number': vlanid,
    'description': description})
  IPAMvlans = ipam.get('/vlan/', {
    'filter_by': 'domainId',
    'filter_value': 3,
  })
  for IPAMVlan in IPAMvlans:
    current = int(IPAMVlan['number'])
    if current == vlanid:
        ipamid = IPAMVlan['vlanId']
  return ipamid

def CreateExternVlan(vlanid, name, description, customcb3, customl3):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipamurl = config.get('IPAM', 'url')
  ipampassword = config.get('IPAM', 'ipampassword')


  ipam = PhpIpamClient(
    url=ipamurl,
    app_id='network',
    username=ipamuser,
    ssl_verify=False,
    password=ipampassword,
    user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.post('/vlan/', {
    'domainId': 3,
    'name': name,
    'number': vlanid,
    'description': description,
    'custom_CB3': customcb3,
    'custom_L3': customl3,
    'custom_VRF': 'Extern'})
  IPAMvlans = ipam.get('/vlan/', {
    'filter_by': 'domainId',
    'filter_value': 3,
  })
  for IPAMVlan in IPAMvlans:
    current = int(IPAMVlan['number'])
    if current == vlanid:
        ipamid = IPAMVlan['vlanId']
  return ipamid

def CreateServiceVlan(vlanid, name, description):
  """[summary]

  Args:
      vlanid (int): VLAN ID
      name (str): customer123 customer ID
      description (str): Kunden Name

  Returns:
      [type]: [description]
  """
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipamurl = config.get('IPAM', 'url')
  ipampassword = config.get('IPAM', 'ipampassword')


  ipam = PhpIpamClient(
    url=ipamurl,
    app_id='network',
    username=ipamuser,
    ssl_verify=False,
    password=ipampassword,
    user_agent='myapiclient', # custom user-agent header
  )
  IPAMvlans = ipam.post('/vlan/', {
    'domainId': 3,
    'name': name,
    'number': vlanid,
    'description': description,
    'custom_CB3': 1,
    'custom_L3': 1,
    'custom_VRF': 'Service'})
  IPAMvlans = ipam.get('/vlan/', {
    'filter_by': 'domainId',
    'filter_value': 3,
  })
  for IPAMVlan in IPAMvlans:
    current = int(IPAMVlan['number'])
    if current == vlanid:
        ipamid = IPAMVlan['vlanId']
  return ipamid

def NextFreeVlan(minimum, maximum):
  """[summary]

  Args:
      minimum (int): Freies VLAN suchen groesser als
      maximum (int): Freies VLAN suchen kleiner als

  Returns:
      int: Freie VLAN Nummer
  """
  config = ConfigParser()

  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipamurl = config.get('IPAM', 'url')
  ipampassword = config.get('IPAM', 'ipampassword')


  ipam = PhpIpamClient(
      url=ipamurl,
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMvlans = ipam.get('/vlan/', {
      'filter_by': 'domainId',
      'filter_value': 3,
  })
  networks = []

  for IPAMVlan in IPAMvlans:
    if int(IPAMVlan['number']) > minimum and int(IPAMVlan['number']) < 4000:
      networks.append(int(IPAMVlan['number']))


  # Sortieren
  networks = sorted(networks)
  bluevlan = 0
  # Lücke finden
  last = minimum
  for network in networks:
    if network > minimum and network < 4000:
      if (last + 1) == network:
        last = network
      elif (last + 1) != network:
        bluevlan = last + 1
  if bluevlan == 0:
    return None
  return bluevlan

def CreateCustomerSlash56(description):
  config = ConfigParser()
  config.read('C:/Temp/Git/Fortinet/Fortigate/settings.ini')

  ipamuser = config.get('IPAM', 'ipamuser')
  ipampassword = config.get('IPAM', 'ipampassword')

  ipam = PhpIpamClient(
      url='https://ipam.consinto.com',
      app_id='network',
      username=ipamuser,
      ssl_verify=False,
      password=ipampassword,
      user_agent='myapiclient', # custom user-agent header
  )

  IPAMv6subnets = ipam.get('/subnets/432/first_subnet/56/')
  # Das neue Netz wird angelegt
  IPAMvlans = ipam.post('/subnets/432/first_subnet/56/', {
      'description': description
  })
  searchstring = '/subnets/cidr/' + IPAMv6subnets + '/'
  v6Info = ipam.get(searchstring, {
      'filter_by': 'sectionId',
      'filter_value': 2,
  })
  v6Info = json.dumps(v6Info[0])
  v6Info = json.loads(v6Info)
  bluev6Id = v6Info['id']
  return bluev6Id


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()